from flask import Flask, render_template, request, jsonify
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Ruta para renderizar la página principal
@app.route('/')
def index():
    return render_template('index.html')

# Función para generar la CURP
def generar_curp(datos):
    # Obtener datos del formulario
    nombre = datos['nombre'].upper()
    apellido1 = datos['apellido1'].upper()
    apellido2 = datos['apellido2'].upper()
    fecha_nacimiento = datos['fechaNacimiento']
    sexo = datos['sexo']
    entidad_nacimiento = datos['entidad'][:2]  # Tomar solo las primeras dos letras de la entidad

    # Obtener año, mes y día de nacimiento
    año_nacimiento, mes_nacimiento, dia_nacimiento = fecha_nacimiento.split('-')

    # Obtener primera consonante del primer apellido
    primera_consonante_apellido1 = ''
    for letra in apellido1[1:]:  # Empezar desde la segunda letra
        if letra not in ['A', 'E', 'I', 'O', 'U']:
            primera_consonante_apellido1 = letra
            break

    # Obtener primera consonante del segundo apellido
    primera_consonante_apellido2 = ''
    for letra in apellido2[1:]:  # Empezar desde la segunda letra
        if letra not in ['A', 'E', 'I', 'O', 'U']:
            primera_consonante_apellido2 = letra
            break

    # Obtener primera consonante del nombre
    primera_consonante_nombre = ''
    for letra in nombre[1:]:  # Empezar desde la segunda letra
        if letra not in ['A', 'E', 'I', 'O', 'U']:
            primera_consonante_nombre = letra
            break

    # Formar la CURP
    curp = f"{apellido1[:2]}{apellido2[0]}{nombre[0]}{año_nacimiento[2:]}{mes_nacimiento}{dia_nacimiento}{sexo}{entidad_nacimiento}{primera_consonante_apellido1}{primera_consonante_apellido2}{primera_consonante_nombre}"

    return curp


# Función para guardar la CURP en un archivo Excel
def guardar_curp_en_excel(curp):
    try:
        # Intenta cargar el archivo Excel existente
        wb = load_workbook('curps.xlsx')
        ws = wb.active
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        wb = Workbook()
        ws = wb.active
        # Agrega encabezados si es un archivo nuevo
        ws.append(['CURP'])

    # Agrega la CURP en una nueva fila
    ws.append([curp])

    # Guarda el archivo Excel
    wb.save('curps.xlsx')

# Ruta para procesar los datos del formulario y generar la CURP
@app.route('/generar_curp', methods=['POST'])
def generar_y_guardar_curp():
    datos = request.form
    curp_generada = generar_curp(datos)
    guardar_curp_en_excel(curp_generada)
    return jsonify({"curp": curp_generada})

if __name__ == '__main__':
    app.run(debug=True)
